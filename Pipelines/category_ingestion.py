import csv
import docx
import os
import sys

# Add root directory to python path to import modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from Agent.category_graph import CategoryGraph
from sql_script import Database

def process_row(graph: CategoryGraph, folder_name: str, row_data: list, row_index: int, document_name: str, question_cols: int, response_cols: int):
    """
    Parses the row_data according to layout and runs it through the LangGraph categorizer.
    """
    min_cols = 1 + question_cols + response_cols
    if not isinstance(row_data, list) or len(row_data) < min_cols:
        print(f"[WARN] Skipping row {row_index} in {document_name}: Invalid format or missing columns.")
        return

    # Check for header
    try:
        idx = int(str(row_data[0]).strip())
    except ValueError:
        if str(row_data[0]).strip().lower() == 'index':
            print("Skipping header row.")
            return

    # Dynamic question columns
    q_start = 1
    q_end = q_start + question_cols
    question_parts = [str(row_data[i]).strip() for i in range(q_start, q_end) if str(row_data[i]).strip()]
    question_text = "\n".join(question_parts)
    
    # Dynamic response columns
    r_start = q_end
    r_end = r_start + response_cols
    response_parts = [str(row_data[i]).strip() for i in range(r_start, min(r_end, len(row_data))) if str(row_data[i]).strip()]
    response_text = "\n".join(response_parts)

    if not question_text:
        print(f"[WARN] Skipping row {row_index}: No question text extracted.")
        return

    print(f"\nProcessing Row {row_index} ...")
    
    # Run through the LangGraph workflow
    result = graph.process_ingestion_row(
        folder_name=folder_name,
        question=question_text,
        response=response_text,
        document_name=document_name,
        row_index=row_index
    )

    assigned_cats = result.get("assigned_categories", [])
    cat_names = [c.category_name for c in assigned_cats]
    print(f" -> Assigned to: {', '.join(cat_names)}")


def ingest_csv(graph: CategoryGraph, folder_name: str, file_path: str, question_cols: int, response_cols: int):
    doc_name = os.path.basename(file_path)
    print(f"Processing CSV: {file_path}")
    try:
        with open(file_path, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            row_idx = 0
            for row in reader:
                process_row(graph, folder_name, row, row_idx, doc_name, question_cols, response_cols)
                row_idx += 1
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

def ingest_docx(graph: CategoryGraph, folder_name: str, file_path: str, question_cols: int, response_cols: int):
    doc_name = os.path.basename(file_path)
    print(f"Processing DOCX: {file_path}")
    try:
        doc = docx.Document(file_path)
        row_idx = 0
        for table in doc.tables:
            for row in table.rows:
                row_data = [cell.text.strip() for cell in row.cells]
                process_row(graph, folder_name, row_data, row_idx, doc_name, question_cols, response_cols)
                row_idx += 1
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

def ingest_xlsx(graph: CategoryGraph, folder_name: str, file_path: str, question_cols: int, response_cols: int, sheet_configs=None):
    import openpyxl
    doc_name = os.path.basename(file_path)
    print(f"Processing XLSX: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if sheet_configs and sheet in sheet_configs:
                s_qcols, s_rcols = sheet_configs[sheet]
            else:
                s_qcols, s_rcols = question_cols, response_cols
            
            print(f"  Sheet: {sheet}")
            row_idx = 0
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                process_row(graph, folder_name, row_data, row_idx, doc_name, s_qcols, s_rcols)
                row_idx += 1
        wb.close()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")


def process_file(doc_name: str, folder_name: str, question_cols: int, response_cols: int, sheet_configs=None):
    file_path = os.path.join("downloaded_documents", doc_name)
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print("Initializing LangGraph")
    graph = CategoryGraph()
    folder_cats = graph.categories.get(folder_name, [])
    print(f"Preloaded {len(folder_cats)} categories for {folder_name}.")

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    if ext == '.csv':
        ingest_csv(graph, folder_name, file_path, question_cols, response_cols)
    elif ext == '.docx':
        ingest_docx(graph, folder_name, file_path, question_cols, response_cols)
    elif ext == '.xlsx':
        ingest_xlsx(graph, folder_name, file_path, question_cols, response_cols, sheet_configs)
    else:
        print(f"Unsupported file format: {ext}")

if __name__ == "__main__":
    print("\n--- Category-based Ingestion ---")
    doc_name = input("Enter the document name (e.g., file.csv, file.xlsx): ").strip()
    if not doc_name:
        print("No document provided.")
        sys.exit(0)

    folder_name = input("Enter the folder name / domain (e.g., AML, ALM, FM): ").strip()
    if not folder_name:
        print("Folder name is required for category isolation.")
        sys.exit(0)

    _, ext = os.path.splitext(doc_name)
    sheet_configs = None

    if ext.lower() == '.xlsx':
        import openpyxl
        file_path = os.path.join("downloaded_documents", doc_name)
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            
            print(f"\nFound {len(sheets)} sheet(s): {', '.join(sheets)}")
            sheet_configs = {}
            for sheet in sheets:
                print(f"\n--- Sheet: '{sheet}' ---")
                q_input = input(f"  Number of question columns (default 2): ").strip()
                q_cols = int(q_input) if q_input else 2
                r_input = input(f"  Number of response columns (default 1): ").strip()
                r_cols = int(r_input) if r_input else 1
                sheet_configs[sheet] = (q_cols, r_cols)
        except Exception as e:
            print(f"Falling back to default layout. ({e})")
            sheet_configs = None

        question_cols = 2
        response_cols = 1
    else:
        q_cols_input = input("Number of columns for question (default 2): ").strip()
        question_cols = int(q_cols_input) if q_cols_input else 2
        
        r_cols_input = input("Number of columns for response (default 1): ").strip()
        response_cols = int(r_cols_input) if r_cols_input else 1

    process_file(
        doc_name=doc_name,
        folder_name=folder_name,
        question_cols=question_cols,
        response_cols=response_cols,
        sheet_configs=sheet_configs
    )
    print("\n[OK] Processing complete.")
