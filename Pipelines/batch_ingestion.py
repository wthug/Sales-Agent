import csv
import docx
import os
import sys

# Add root directory to python path to import sql_script
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from sql_script import Database
from langchain_openai import OpenAIEmbeddings, ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage
from dotenv import load_dotenv

load_dotenv()
open_api_key = os.getenv("OPENAI_API_KEY")

def get_embeddings(input_text: str) -> list:
    try: 
        embeddings = OpenAIEmbeddings(
            api_key=open_api_key,
            model="text-embedding-3-small",
            dimensions=384
        )
        embedding_vector = embeddings.embed_query(input_text)
        return embedding_vector
    except Exception as e:
        # print(f"Error generating embeddings: {e}")
        return []

def process_row(row_data, row_index=None, document_name=None, document_id=None, folder_name=None, document_sharepoint_url=None, question_cols=2, response_cols=1):
    """
    Function called for every row extracted from the file.
    
    Column layout (0-indexed):
      col 0           → row index
      col 1 .. question_cols  → concatenated to form question_text
      col (1+question_cols) .. (1+question_cols+response_cols-1) → concatenated to form response_text
    """
    print(row_data)
    
    min_cols = 1 + question_cols + response_cols
    response_text = ""
    
    if isinstance(row_data, list) and len(row_data) >= min_cols:
        try:
            row_index = int(str(row_data[0]).strip())
        except ValueError:
            if str(row_data[0]).strip().lower() == 'index':
                print("Skipping header row.")
                return
        
        # Dynamic question columns
        q_start = 1
        q_end = q_start + question_cols
        question_parts = [str(row_data[i]).strip() for i in range(q_start, q_end) if str(row_data[i]).strip()]
        question_text = "\n".join(question_parts)
        
        # Dynamic response columns
        r_start = q_end
        r_end = r_start + response_cols
        response_parts = [str(row_data[i]).strip() for i in range(r_start, min(r_end, len(row_data))) if str(row_data[i]).strip()]
        response_text = "\n".join(response_parts)
    else:
        question_text = " | ".join([str(cell) for cell in row_data if cell]) if isinstance(row_data, list) else str(row_data)
    
    if not question_text:
        print(f"[WARN] Skipping row {row_index} in document '{document_name}': No question text extracted.")
        return

    embeddings = get_embeddings(question_text)
    if not embeddings:
        print(f"[WARN] Skipping row {row_index} in document '{document_name}': Failed to generate embeddings.")
        return

    conn = Database.get_connection()
    if conn:
        try:
            with conn.cursor() as cur:
                # Check for similar question (cosine similarity >= 0.90)
                cur.execute('''
                    SELECT que_id, response,folder_name, (1 - (embeddings <=> %s::vector)) as similarity 
                    FROM batch_questions 
                    WHERE folder_name = %s
                    ORDER BY embeddings <=> %s::vector 
                    LIMIT 1;
                ''', (embeddings, folder_name, embeddings))
                
                result = cur.fetchone()
                if result and result[3] is not None and result[3] >= 0.90:
                    que_id = result[0]
                    existing_response = result[1]
                    similarity = result[3]
                    print(f"Found similar question (Accuracy: {similarity:.2f}) for row {row_index}. Merging responses...")
                    
                    llm = ChatOpenAI(model="gpt-4o-mini", api_key=open_api_key)
                    prompt_text = f"Existing Response:\n{existing_response}\n\nNew Response:\n{response_text}"
                    messages = [
                        SystemMessage(content="You are an expert at summarizing and merging information. Please summarize the two responses into a single comprehensive response without losing critical details. Output only the summarized response."),
                        HumanMessage(content=prompt_text)
                    ]
                    summarized_response = llm.invoke(messages).content
                    
                    cur.execute('''
                        UPDATE batch_questions 
                        SET response = %s 
                        WHERE que_id = %s
                    ''', (summarized_response, que_id))
                    print(f"Updated row {row_index} by merging with existing question (que_id: {que_id}).\n\n")
                else:
                    cur.execute('''
                        INSERT INTO batch_questions (
                            document_id, question, response, embeddings, folder_name, row_index, document_name, document_sharepoint_url
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (document_id, question_text, response_text, embeddings, folder_name, row_index, document_name, document_sharepoint_url))
                    print(f"Saved row {row_index} of {document_name} to database.\n\n")
                
                conn.commit()
        except Exception as e:
            print(f"Error processing row {row_index} to database: {e}\n\n")
            conn.rollback()
        finally:
            Database.put_connection(conn)

def ingest_docx(file_path, document_id=None, folder_name=None, document_sharepoint_url=None, question_cols=2, response_cols=1):
    """
    Iterates through all tables in a DOCX file and calls process_row for each row.
    Ignores all text outside of tables.
    """
    print(f"Processing DOCX: {file_path}")
    doc_name = os.path.basename(file_path)
    try:
        doc = docx.Document(file_path)
        row_idx = 0
        for table in doc.tables:
            for row in table.rows:
                row_data = [cell.text.strip() for cell in row.cells]
                process_row(row_data, row_index=row_idx, document_name=doc_name, document_id=document_id, folder_name=folder_name, document_sharepoint_url=document_sharepoint_url, question_cols=question_cols, response_cols=response_cols)
                row_idx += 1
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

def ingest_csv(file_path, document_id=None, folder_name=None, document_sharepoint_url=None, question_cols=2, response_cols=1):
    """
    Iterates through all rows in a CSV file and calls process_row for each row.
    """
    print(f"Processing CSV: {file_path}")
    doc_name = os.path.basename(file_path)
    try:
        with open(file_path, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            row_idx = 0
            for row in reader:
                process_row(row, row_index=row_idx, document_name=doc_name, document_id=document_id, folder_name=folder_name, document_sharepoint_url=document_sharepoint_url, question_cols=question_cols, response_cols=response_cols)
                row_idx += 1
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

def ingest_xlsx(file_path, document_id=None, folder_name=None, document_sharepoint_url=None, question_cols=2, response_cols=1, sheet_configs=None):
    """
    Iterates through all rows in an XLSX file and calls process_row for each row.
    
    sheet_configs: optional dict mapping sheet name -> (question_cols, response_cols)
                   If a sheet is not in the dict, the top-level question_cols/response_cols are used.
    """
    import openpyxl
    print(f"Processing XLSX: {file_path}")
    doc_name = os.path.basename(file_path)
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Per-sheet column config
            if sheet_configs and sheet in sheet_configs:
                s_qcols, s_rcols = sheet_configs[sheet]
            else:
                s_qcols, s_rcols = question_cols, response_cols
            
            print(f"  Sheet: {sheet}  (question_cols={s_qcols}, response_cols={s_rcols})")
            row_idx = 0
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                process_row(row_data, row_index=row_idx, document_name=doc_name, document_id=document_id, folder_name=folder_name, document_sharepoint_url=document_sharepoint_url, question_cols=s_qcols, response_cols=s_rcols)
                row_idx += 1
        wb.close()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

def process_file(doc_name, folder_name=None, document_sharepoint_url=None, question_cols=2, response_cols=1, sheet_configs=None):
    """
    Determines the file type and routes to the appropriate ingestion function.
    """
    file_path = os.path.join("downloaded_documents", doc_name)
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    # Insert document into DB with status 'false' and get document_id
    conn = Database.get_connection()
    document_id = None
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO documents (file_name, sharepoint_url, folder_name, processing_type)
                    VALUES (%s, %s, %s, 'batch')
                    RETURNING document_id;
                """, (doc_name, document_sharepoint_url, folder_name))
                document_id = cur.fetchone()[0]
                conn.commit()
                print(f"Inserted document '{doc_name}' with ID: {document_id}")
        except Exception as e:
            print(f"Error inserting document into DB: {e}")
            conn.rollback()
        finally:
            Database.put_connection(conn)
            
    if not document_id:
        print("Failed to get document_id, skipping processing.")
        return

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    if ext == '.docx':
        ingest_docx(file_path, document_id, folder_name=folder_name, document_sharepoint_url=document_sharepoint_url, question_cols=question_cols, response_cols=response_cols)
    elif ext == '.csv':
        ingest_csv(file_path, document_id, folder_name=folder_name, document_sharepoint_url=document_sharepoint_url, question_cols=question_cols, response_cols=response_cols)
    elif ext == '.xlsx':
        ingest_xlsx(file_path, document_id, folder_name=folder_name, document_sharepoint_url=document_sharepoint_url, question_cols=question_cols, response_cols=response_cols, sheet_configs=sheet_configs)
    else:
        print(f"Unsupported file format: {ext}")

    # Update status to 'true' after processing
    conn = Database.get_connection()
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE documents SET ingestion_status = 'successful' WHERE document_id = %s;
                """, (document_id,))
                conn.commit()
                print(f"Updated ingestion_status to 'successful' for document '{doc_name}'")
        except Exception as e:
            print(f"Error updating document status: {e}")
            conn.rollback()
        finally:
            Database.put_connection(conn)
            
    return document_id

if __name__ == "__main__":
    doc_name = input("Enter the document name (e.g., file.docx, file.csv, file.xlsx): ").strip()
    folder_name = input("Enter the folder name (Optional): ").strip()
    sharepoint_url = input("Enter the SharePoint URL (Optional): ").strip()
    
    _, ext = os.path.splitext(doc_name)
    sheet_configs = None
    
    if ext.lower() == '.xlsx':
        # For XLSX: prompt per-sheet column configuration
        import openpyxl
        file_path = os.path.join("downloaded_documents", doc_name)
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            
            print(f"\nFound {len(sheets)} sheet(s): {', '.join(sheets)}")
            sheet_configs = {}
            for sheet in sheets:
                print(f"\n--- Sheet: '{sheet}' ---")
                q_input = input(f"  Number of question columns (default 2): ").strip()
                q_cols = int(q_input) if q_input else 2
                r_input = input(f"  Number of response columns (default 1): ").strip()
                r_cols = int(r_input) if r_input else 1
                sheet_configs[sheet] = (q_cols, r_cols)
                print(f"  Layout: [index] + [{q_cols} question col(s)] + [{r_cols} response col(s)]")
        except Exception as e:
            print(f"Could not read sheets from file: {e}")
            print("Falling back to default column layout for all sheets.")
            sheet_configs = None
        
        # Defaults (used as fallback if sheet_configs is None)
        question_cols = 2
        response_cols = 1
    else:
        # For CSV/DOCX: single column config
        q_cols_input = input("Number of columns for question (default 2): ").strip()
        question_cols = int(q_cols_input) if q_cols_input else 2
        
        r_cols_input = input("Number of columns for response (default 1): ").strip()
        response_cols = int(r_cols_input) if r_cols_input else 1
        
        print(f"\nColumn layout: [index] + [{question_cols} question col(s)] + [{response_cols} response col(s)]")
    
    if doc_name:
        process_file(
            doc_name, 
            folder_name=folder_name if folder_name else None, 
            document_sharepoint_url=sharepoint_url if sharepoint_url else None,
            question_cols=question_cols,
            response_cols=response_cols,
            sheet_configs=sheet_configs
        )
    else:
        print("No document name provided.")

